import os
from dotenv import load_dotenv
from langchain_groq import ChatGroq
from pathlib import Path
from pypdf import PdfReader  # ← added for PDF support
import glob
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Define character & rules once at the top
SYSTEM_PROMPT = """
You are DataBot – accurate enterprise data extraction assistant.

Rules:
- Output ONLY valid JSON. No other text, no markdown, no explanations.
- Use snake_case keys.
- Infer meaningful field names (employee_name, vendor_name, invoice_number, po_number, total_amount, currency, invoice_date, billing_period, hours_worked, project_code, line_items, etc.)
- Include line_items as array when present.
- Always return:
  - status: "complete" | "partial" | "needs_review" | "invalid"
  - document_type: "invoice" | "timesheet" | "receipt" | "unknown"
  - extracted_fields: object with fields
  - validation_issues: array of {field, severity, message}
  - overall_confidence: number 0.0–1.0

Example output:
{
  "status": "partial",
  "document_type": "invoice",
  "extracted_fields": {
    "vendor_name": "Beta Solutions GmbH",
    "invoice_number": "INV-78421",
    "po_number": "PO-998877",
    "total_amount": 2525.50,
    "currency": "EUR",
    "invoice_date": "2026-02-09"
  },
  "validation_issues": [
    {"field": "total_amount", "severity": "warning", "message": "Line sum matches total"}
  ],
  "overall_confidence": 0.92
}
"""

load_dotenv()

# LLM Setup
llm = ChatGroq(
    model="llama-3.1-8b-instant",  # fast & reliable on free tier
    temperature=0.3,
    groq_api_key=os.getenv("GROQ_API_KEY")
)

# Simple memory storage (no embedding issues)
memories = []
last_input_data = None  # Store last input for "try again"

def get_memories(query: str) -> str:
    if memories:
        return f"Relevant previous knowledge:\n" + "\n".join(memories) + "\n"
    return ""

def save_memory(content: str) -> None:
    memories.append(content)
    print("✓ Saved to memory.")

def save_to_excel(response_text: str, excel_file: str = "output.xlsx") -> None:
    """
    Parse DataBot structured output and save to Excel.
    Handles Records, Record, and Validation formats.
    """
    try:
        # Check if file exists; if not, create new workbook
        file_path = Path(excel_file)
        if file_path.exists():
            wb = load_workbook(excel_file)
            ws = wb.active
            next_row = ws.max_row + 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            next_row = 1
        
        lines = response_text.strip().split('\n')
        
        # Check format and parse accordingly
        if lines[0].startswith("Records:"):
            # Multiple records format: numbered list
            current_row = next_row + 1 if next_row > 1 else 2  # Leave space for headers
            headers_set = False
            
            for line in lines[1:]:
                if line.startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.')):
                    # Extract record number and data
                    record_data = line.split('.', 1)[1].strip() if '.' in line else line
                    fields = [f.strip() for f in record_data.split('|')]
                    
                    if not headers_set and next_row == 1:
                        # Set headers from field names
                        header_row = []
                        for field in fields:
                            if ':' in field:
                                header = field.split(':')[0].strip()
                                header_row.append(header)
                        if header_row:
                            for col, header in enumerate(header_row, 1):
                                ws.cell(row=1, column=col, value=header)
                            current_row = 2
                        headers_set = True
                    
                    # Add data row
                    for col, field in enumerate(fields, 1):
                        value = field.split(':', 1)[1].strip() if ':' in field else field
                        ws.cell(row=current_row, column=col, value=value)
                    current_row += 1
        
        elif lines[0].startswith("Record:"):
            # Single record format
            headers = []
            values = []
            
            for line in lines[1:]:
                if line.startswith("-"):
                    parts = line.lstrip("- ").split(':', 1)
                    if len(parts) == 2:
                        headers.append(parts[0].strip())
                        values.append(parts[1].strip())
            
            # Add headers if not already present
            if next_row == 1:
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                current_row = 2
            else:
                current_row = next_row
            
            # Add values
            for col, value in enumerate(values, 1):
                ws.cell(row=current_row, column=col, value=value)
        
        wb.save(excel_file)
        print(f"✓ Data saved to {excel_file}")
    
    except Exception as e:
        print(f"Error saving to Excel: {e}")

def process_data(input_data: str, user_input: str) -> None:
    """Process input data with DataBot and store for retries."""
    global last_input_data
    last_input_data = input_data
    
    memories_str = get_memories(user_input)
    prompt = f"""{SYSTEM_PROMPT}

{memories_str if memories_str else "you remember everything i taught you."}

Input data to process:
{input_data}

Extract and structure the data according to your rules.
Output ONLY in your defined structured format (Record / Records / Validation).
"""
    
    print("Thinking...\n")
    try:
        response = llm.invoke(prompt)
        output = response.content.strip()
        print(f"DataBot:\n{output}\n")
        
        # Save to Excel if it's record data (not validation)
        if "Record:" in output or "Records:" in output:
            save_to_excel(output)
    except Exception as e:
        print(f"Error calling Groq: {e}")

# ────────────────────────────────────────────────
#   NEW FUNCTION: Extract text from files (txt + pdf)
# ────────────────────────────────────────────────
def extract_text_from_file(file_path: str) -> str:
    """
    Reads text from a file (supports .txt and .pdf).
    Returns the extracted text or an error message.
    """
    file_path = str(Path(file_path).resolve())  # make absolute path
    ext = Path(file_path).suffix.lower()

    if ext in [".txt", ".eml"]:
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception as e:
            return f"[Error reading text file: {e}]"

    elif ext == ".pdf":
        try:
            reader = PdfReader(file_path)
            text = ""
            for page in reader.pages:
                page_text = page.extract_text() or ""
                text += page_text + "\n"
            if text.strip():
                return text.strip()
            else:
                return "[PDF has no extractable text — possibly scanned/image-based. OCR needed.]"
        except Exception as e:
            return f"[Error reading PDF: {e}]"

    else:
        return f"[Unsupported file type: {ext}. Supported: .txt, .pdf]"

def find_files_by_pattern(pattern: str, search_dir: str = ".") -> list:
    """
    Search for files matching a pattern (case-insensitive).
    Returns a list of matching file paths.
    """
    search_path = Path(search_dir).resolve()
    if not search_path.exists():
        return []
    
    # Search for files with pattern in name (case-insensitive)
    pattern_lower = pattern.lower()
    matches = []
    
    for file_path in search_path.rglob("*"):
        if file_path.is_file() and pattern_lower in file_path.name.lower():
            matches.append(str(file_path))
    
    return sorted(matches)

def run_agent():
    print("\n=== DataBot — Data Entry Specialist ===")
    print("Commands:")
    print("  teach: <rule>                → teach a formatting or business rule")
    print("  file: <path/to/file.pdf>     → process a specific PDF or txt file")
    print("  search <pattern>             → search for & process file with pattern in name (e.g., 'search invoice')")
    print("  try again / retry            → reprocess last input with updated rules")
    print("  Any other text               → treat as direct data entry input")
    print("  exit / quit                  → stop\n")

    while True:
        user_input = input("\nYou: ").strip()
        if user_input.lower() in ['exit', 'quit', 'q']:
            print("Goodbye!")
            break

        # Teach rule
        if user_input.lower().startswith(('teach:', 'remember:', 'note that')):
            content = user_input.split(':', 1)[1].strip() if ':' in user_input else user_input
            save_memory(content)
            continue

        # Try again with updated rules
        if user_input.lower() in ['try again', 'retry', 'try', 'again']:
            if last_input_data is None:
                print("No previous task to retry. Please provide data first.\n")
                continue
            print("Retrying with updated rules...\n")
            process_data(last_input_data, user_input)
            continue

        # Search for file by pattern
        if user_input.lower().startswith(("search:", "search ")):
            # Accept both "search: pattern" and "search pattern"
            pattern = user_input.split(None, 1)[1].strip() if ' ' in user_input.split(':', 1)[0] else user_input.split(":", 1)[1].strip()
            files = find_files_by_pattern(pattern)
            
            if not files:
                print(f"No files found matching '{pattern}'")
                continue
            
            print(f"Found {len(files)} file(s):")
            for i, f in enumerate(files, 1):
                print(f"  {i}. {f}")
            
            if len(files) == 1:
                file_path = files[0]
                print(f"\nProcessing: {file_path}")
            else:
                # If multiple files found, ask user to choose
                choice = input(f"Which file? (1-{len(files)}, or 'all'): ").strip()
                if choice.lower() == 'all':
                    # Process all files
                    for file_path in files:
                        print(f"\n--- Processing: {file_path} ---")
                        file_content = extract_text_from_file(file_path)
                        process_data(file_content, user_input)
                    continue
                else:
                    try:
                        idx = int(choice) - 1
                        if 0 <= idx < len(files):
                            file_path = files[idx]
                        else:
                            print("Invalid choice")
                            continue
                    except ValueError:
                        print("Invalid input")
                        continue
            
            file_content = extract_text_from_file(file_path)
            print("File content extracted. Processing...\n")
            process_data(file_content, user_input)
            continue

        # File processing
        if user_input.lower().startswith("file:"):
            file_path = user_input.split(":", 1)[1].strip()
            if not os.path.exists(file_path):
                print(f"File not found: {file_path}")
                continue

            print(f"Reading file: {file_path}")
            file_content = extract_text_from_file(file_path)
            print("File content extracted. Processing...\n")
            process_data(file_content, user_input)
            continue

        # Direct data entry input
        process_data(user_input, user_input)

if __name__ == "__main__":
    if not os.getenv("GROQ_API_KEY"):
        print("ERROR: Add GROQ_API_KEY to .env file")
        exit(1)
    run_agent()